import json
import os
import re
from datetime import datetime
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

import config
from db import get_conn, get_or_create_agent

# Manifest column header -> internal field name. Matches the sample manifest
# (agent_name, call_number, cti_call_number, call_start_time, call_service_name,
#  call_lead_id, call_end_type_name, call_talk_duration, call_trunk_duration, Recordings).
COLUMN_MAP = {
    "agent_name": "agent_name",
    "call_number": "call_number",
    "cti_call_number": "cti_call_number",
    "call_start_time": "call_start_time",
    "call_service_name": "call_service_name",
    "call_lead_id": "call_lead_id",
    "call_end_type_name": "call_end_type_name",
    "call_talk_duration": "call_talk_duration",
    "call_trunk_duration": "call_trunk_duration",
    "recordings": "recording",
}

DATE_FORMATS = [
    "%m-%d-%Y %H:%M",
    "%d-%m-%Y %H:%M",
    "%Y-%m-%d %H:%M",
    "%m/%d/%Y %H:%M",
    "%d-%b-%y %H:%M",
]


def parse_call_date(raw_value):
    """Best-effort parse of call_start_time into a YYYY-MM-DD string for filtering."""
    if raw_value is None:
        return None
    if isinstance(raw_value, datetime):
        return raw_value.strftime("%Y-%m-%d")
    text = str(raw_value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def extract_recording_base(cell_value, hyperlink_target=None):
    """
    Pulls the shared base filename out of a manifest 'Recordings' cell, e.g.
      "7002_08320801862_17-Jun-26-18-11-32.wav16"  ->  "7002_08320801862_17-Jun-26-18-11-32"
    Works whether the cell holds a bare filename, a local path, or a full URL/hyperlink.
    """
    source = hyperlink_target or cell_value
    if not source:
        return None
    source = str(source).strip()
    # take just the filename if it's a path or URL
    parsed = urlparse(source)
    filename = os.path.basename(parsed.path) if parsed.scheme else os.path.basename(source)
    # strip extension
    base = re.sub(r"\.wav16?$", "", filename, flags=re.IGNORECASE)
    # strip any existing -IN / -OUT suffix so we always get the shared base
    base = re.sub(r"-(IN|OUT)$", "", base, flags=re.IGNORECASE)
    return base or None


# def build_leg_paths(recording_base, agent_name, recording_base_url):
#     base_url = (recording_base_url or config.RECORDINGS_BASE_URL).rstrip("/")
#     folder = requests.utils.quote(agent_name or "")
#     rx_path = f"{base_url}/{folder}/{recording_base}-IN.wav16"
#     tx_path = f"{base_url}/{folder}/{recording_base}-OUT.wav16"
#     return rx_path, tx_path
def build_leg_paths(recording_base, agent_name, recording_base_url):
    # base_url = (recording_base_url or config.RECORDINGS_BASE_URL).rstrip("/")
    # folder = requests.utils.quote(agent_name or "")
    # # rx_path = f"{base_url}/{folder}/{recording_base}-IN.wav16"
    # rx_path = f"{recording_base_url or config.RECORDINGS_BASE_URL}/{recording_base}-IN.wav16"
    # # tx_path = f"{base_url}/{folder}/{recording_base}-OUT.wav16"
    # tx_path = f"{recording_base_url or config.RECORDINGS_BASE_URL}/{recording_base}-OUT.wav16"
    path,ext=os.path.splitext(recording_base)
    rx_path = f"{recording_base_url or config.RECORDINGS_BASE_URL}/{path}-IN{ext}"
    tx_path = f"{recording_base_url or config.RECORDINGS_BASE_URL}/{path}-OUT{ext}"
    return rx_path, tx_path


def parse_manifest_file(file_path):
    """Reads the uploaded xlsx and returns a list of row dicts keyed by COLUMN_MAP values."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        key = str(h).strip().lower().replace(" ", "_")
        if key in COLUMN_MAP:
            header_index[COLUMN_MAP[key]] = i

    rows = []
    for excel_row in ws.iter_rows(min_row=2):
        if all(c.value is None for c in excel_row):
            continue
        record = {}
        for field, idx in header_index.items():
            cell = excel_row[idx]
            if field == "recording":
                record["recording_raw"] = cell.value
                record["recording_hyperlink"] = cell.hyperlink.target if cell.hyperlink else None
            else:
                record[field] = cell.value
        rows.append(record)
    return rows


# def save_manifest(file_path, filename, recording_base_url):
#     """Parses the manifest and inserts one 'pending' call row per manifest row."""
#     rows = parse_manifest_file(file_path)
#     conn = get_conn()
#     cur = conn.execute(
#         "INSERT INTO manifests (filename, recording_base_url, uploaded_at, total_rows, status) "
#         "VALUES (?, ?, ?, ?, 'uploaded')",
#         (filename, recording_base_url, datetime.utcnow().isoformat(), len(rows)),
#     )
#     manifest_id = cur.lastrowid

#     inserted = 0
#     for r in rows:
#         agent_name = (r.get("agent_name") or "Unassigned").strip()
#         agent_id = get_or_create_agent(conn, agent_name)

#         recording_base = extract_recording_base(r.get("recording_raw"), r.get("recording_hyperlink"))
#         rx_path, tx_path = (None, None)
#         if recording_base:
#             rx_path, tx_path = build_leg_paths(recording_base, agent_name, recording_base_url)

#         call_date = parse_call_date(r.get("call_start_time"))

#         conn.execute(
#             """INSERT INTO calls (
#                 manifest_id, agent_id, call_number, cti_call_number, call_start_time, call_date,
#                 call_service_name, call_lead_id, call_end_type_name, call_talk_duration,
#                 call_trunk_duration, recording_base, rx_path, tx_path, status
#             ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'pending')""",
#             (
#                 manifest_id, agent_id,
#                 str(r.get("call_number") or ""), str(r.get("cti_call_number") or ""),
#                 str(r.get("call_start_time") or ""), call_date,
#                 r.get("call_service_name"), str(r.get("call_lead_id") or ""),
#                 r.get("call_end_type_name"), r.get("call_talk_duration"),
#                 r.get("call_trunk_duration"), recording_base, rx_path, tx_path,
#             ),
#         )
#         inserted += 1

#     conn.commit()
#     conn.close()
#     return manifest_id, inserted

def save_manifest(file_path, filename, recording_base_url):
    """Parses the manifest and inserts one 'pending' call row per manifest row."""
    rows = parse_manifest_file(file_path)
    print(rows)
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO manifests (filename, recording_base_url, uploaded_at, total_rows, status) "
        "VALUES (?, ?, ?, ?, 'uploaded')",
        (filename, recording_base_url, datetime.utcnow().isoformat(), len(rows)),
    )
    manifest_id = cur.lastrowid

    inserted = 0
    for r in rows:
        agent_name = (r.get("agent_name") or "Unassigned").strip()
        agent_id = get_or_create_agent(conn, agent_name)
        # print
        recording_base = extract_recording_base(r.get("recording_raw"), r.get("recording_hyperlink"))
        print("****"*12)
        recording_base=r.get("recording_raw")
        print(recording_base,r.get("recording_raw"))
        rx_path, tx_path = (None, None)
        if recording_base:
            rx_path, tx_path = build_leg_paths(recording_base, agent_name, recording_base_url)
        # rx_path = recording_base
        print("****"*12)
        print(rx_path, tx_path) 
        call_date = parse_call_date(r.get("call_start_time"))

        conn.execute(
            """INSERT INTO calls (
                manifest_id, agent_id, call_number, cti_call_number, call_start_time, call_date,
                call_service_name, call_lead_id, call_end_type_name, call_talk_duration,
                call_trunk_duration, recording_base, rx_path, tx_path, status
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'pending')""",
            (
                manifest_id, agent_id,
                str(r.get("call_number") or ""), str(r.get("cti_call_number") or ""),
                str(r.get("call_start_time") or ""), call_date,
                r.get("call_service_name"), str(r.get("call_lead_id") or ""),
                r.get("call_end_type_name"), r.get("call_talk_duration"),
                r.get("call_trunk_duration"), recording_base, rx_path, tx_path,
            ),
        )
        inserted += 1

    conn.commit()
    conn.close()
    return manifest_id, inserted


def call_stt(rx_path, tx_path):
    resp = requests.post(
        config.STT_API_URL,
        json={"rx_path": rx_path, "tx_path": tx_path},
        timeout=config.REQUEST_TIMEOUT_SECONDS,
    )
    resp.raise_for_status()
    return resp.json()


# def call_slm(conversation_text, utterances):
#     resp = requests.post(
#         config.SLM_API_URL,
#         json={"transcript": conversation_text},
#         json={"conversation": conversation_text, "utterances": utterances},
#         timeout=config.REQUEST_TIMEOUT_SECONDS,
#     )
#     resp.raise_for_status()
#     return resp.json()

def call_slm(conversation_text, utterances):
    resp = requests.post(
        config.SLM_API_URL,
        json={"transcript": conversation_text},
        timeout=config.REQUEST_TIMEOUT_SECONDS,
    )
    resp.raise_for_status()    
    return resp.json()


def normalize_score(param_key, raw_score):
    max_points = config.PARAMETER_MAX_POINTS.get(param_key, config.DEFAULT_PARAMETER_MAX_POINTS)
    if max_points <= 0:
        return 0
    # return round(min(raw_score, max_points) / max_points * 100, 1)
    return raw_score



def quality_from_score(score, fatal_error):
    if fatal_error:
        return "poor"
    if score >= 80:
        return "good"
    if score >= 60:
        return "average"
    return "poor"


def sentiment_from_verdict(verdict, fatal_error):
    if fatal_error:
        return "negative"
    v = (verdict or "").lower()
    if v in ("good", "excellent"):
        return "positive" 
    if v in ("poor", "bad", "fail"): 
        return "negative" 
    return "negative" 


def process_call(conn, call_row):
    call_id = call_row["id"]
    if not call_row["rx_path"] or not call_row["tx_path"]:
        conn.execute(
            "UPDATE calls SET status='failed', error_message=? WHERE id=?",
            ("Could not resolve recording IN/OUT paths from manifest.", call_id),
        )
        return False

    # 1. Speech to text
    conn.execute("UPDATE calls SET status='transcribing' WHERE id=?", (call_id,))
    conn.commit()
    try:
        stt_result = call_stt(call_row["rx_path"], call_row["tx_path"])
    except Exception as e:
        conn.execute(
            "UPDATE calls SET status='failed', error_message=? WHERE id=?",
            (f"STT request failed: {e}", call_id),
        )
        return False

    conversation_text = stt_result.get("conversation", "")
    utterances = stt_result.get("utterances", [])
    conn.execute(
        "INSERT INTO transcripts (call_id, conversation_text, utterances_json, raw_json) VALUES (?,?,?,?)",
        (call_id, conversation_text, json.dumps(utterances), json.dumps(stt_result)),
    )

    # 2. SLM QA analysis
    conn.execute("UPDATE calls SET status='analyzing' WHERE id=?", (call_id,))
    conn.commit()
    try:
        slm_result = call_slm(conversation_text, utterances)
    except Exception as e:
        conn.execute(
            "UPDATE calls SET status='failed', error_message=? WHERE id=?",
            (f"SLM analysis request failed: {e}", call_id),
        )
        return False

    analysis = slm_result.get("analysis", slm_result)
    param_data = analysis.get("parameters", {})
    fatal_data = analysis.get("fatal_checks", {})
    total = analysis.get("total") or param_data.get("total") or 0
    summary = analysis.get("summary", "")
    verdict = analysis.get("verdict", "")
    fatal_error = bool(analysis.get("fatal_error", False))

    param_rows = conn.execute("SELECT id, param_key FROM parameters").fetchall()
    param_id_by_key = {p["param_key"]: p["id"] for p in param_rows}

    scores = []
    for key, pid in param_id_by_key.items():
        entry = param_data.get(key)
        if not entry:
            continue
        raw_score = entry.get("score", 0) or 0
        norm_score = normalize_score(key, raw_score)# keeping the score as it is so the ouput looks good 
        # scores.append(norm_score)
        scores.append(raw_score)

        conn.execute(
            """INSERT INTO call_scores
               (call_id, parameter_id, status, reason, evidence, start_time, end_time, raw_score, score)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (
                call_id, pid, int(bool(entry.get("status"))), entry.get("reason", ""),
                entry.get("evidence", ""), entry.get("start_time"), entry.get("end_time"),
                raw_score, norm_score,
            ),
        )

    for key in config.FATAL_CHECKS:
        entry = fatal_data.get(key, {})
        conn.execute(
            """INSERT INTO fatal_checks (call_id, check_key, status, reason, evidence, start_time, end_time)
               VALUES (?,?,?,?,?,?,?)""",
            (
                call_id, key, int(bool(entry.get("status"))), entry.get("reason", ""),
                entry.get("evidence", ""), entry.get("start_time"), entry.get("end_time"),
            ),
        )

    # overall_score = round(sum(scores) / len(scores), 1) if scores else (total or 0)
    overall_score = sum(scores) if scores else (total or 0)# just sum up the scores 

    overall_quality = quality_from_score(overall_score, fatal_error)
    sentiment = sentiment_from_verdict(verdict, fatal_error)

    conn.execute(
        """UPDATE calls SET status='audited', overall_score=?, overall_quality=?, verdict=?,
           summary=?, fatal_error=?, sentiment=? WHERE id=?""",
        (overall_score, overall_quality, verdict, summary, int(fatal_error), sentiment, call_id),
    )
    return True


def process_manifest(manifest_id):
    conn = get_conn()
    conn.execute("UPDATE manifests SET status='processing' WHERE id=?", (manifest_id,))
    conn.commit()

    pending = conn.execute(
        "SELECT * FROM calls WHERE manifest_id = ? AND status = 'pending'", (manifest_id,)
    ).fetchall()

    ok, failed = 0, 0
    for row in pending:
        success = process_call(conn, row)
        conn.commit()
        if success:
            ok += 1
        else:
            failed += 1

    status = "done_with_errors" if failed else "done"
    conn.execute(
        "UPDATE manifests SET status=?, processed_rows=?, failed_rows=? WHERE id=?",
        (status, ok, failed, manifest_id),
    )
    conn.commit()
    conn.close()
    return {"processed": ok, "failed": failed}
